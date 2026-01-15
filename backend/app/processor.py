from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
import re
from typing import Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.table import Table, TableStyleInfo


@dataclass
class TableRef:
    start_col: int
    start_row: int
    end_col: int
    end_row: int


def process_workbook(content: bytes) -> bytes:
    wb = load_workbook(BytesIO(content))
    ws = wb.active

    # Remove columns F:I
    ws.delete_cols(6, 4)

    # Normalize date columns early (before table creation)
    unit_exp_col = find_header_column(ws, "Unit Expiration Date") or 4
    reg_date_col = find_header_column(ws, "Registration Date") or 5
    normalize_date_column(ws, unit_exp_col)
    normalize_date_column(ws, reg_date_col)

    last_row = find_last_row(ws, 1)
    if last_row < 2:
        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    clear_tables(ws)

    table1_ref = f"A1:E{last_row}"
    add_table(ws, "Table1", table1_ref, style_name="TableStyleMedium2")

    unmerge_row(ws, 1)

    # Copy column B to G
    for row in range(1, last_row + 1):
        ws.cell(row=row, column=7).value = ws.cell(row=row, column=2).value

    # Remove duplicates in G (treat header as data)
    remove_duplicates_in_column(ws, 7, 1, last_row)
    last_row_g = find_last_row(ws, 7)

    # Add Count column in H
    ws.cell(row=1, column=8).value = "Count"
    counts = count_values(ws, 2, 1, last_row)
    for row in range(2, last_row_g + 1):
        key = ws.cell(row=row, column=7).value
        ws.cell(row=row, column=8).value = counts.get(key, 0)

    table2_ref = f"G1:H{last_row_g}"
    add_table(ws, "Table2", table2_ref, style_name="TableStyleMedium10")
    sort_table_by_column(ws, table2_ref, sort_col=8, header_row=1, reverse=True)

    # Locate Unit Expiration Date column
    unit_exp_col = find_header_column(ws, "Unit Expiration Date") or unit_exp_col

    # Insert Quarter column after Unit Expiration Date
    insert_at = unit_exp_col + 1
    ws.insert_cols(insert_at)
    shift_table_columns(ws, insert_at, 1)
    if reg_date_col >= insert_at:
        reg_date_col += 1

    quarter_col = insert_at
    ws.cell(row=1, column=quarter_col).value = "Quarter"
    for row in range(2, find_last_row(ws, 1) + 1):
        cell = ws.cell(row=row, column=unit_exp_col)
        ws.cell(row=row, column=quarter_col).value = date_to_quarter(cell.value)

    # Update Table1 range to include new Quarter column
    table1_ref = expand_table_range(table1_ref, insert_at)
    update_table_ref(ws, "Table1", table1_ref)

    # Quarter_Count output to K:L
    quarter_counts = compute_quarter_counts(ws, quarter_col)
    write_quarter_counts(ws, quarter_counts, start_col=11, start_row=1)

    # Sort Table1 by Unit Expiration Date ascending after all column changes
    table1_ref = get_table_ref(ws, "Table1") or table1_ref
    sort_table_by_date(ws, table1_ref, unit_exp_col)

    # Build header row
    ws.insert_rows(1)
    shift_table_rows(ws, 1, 1)
    build_asset_report_header(ws)
    update_table2_count_formulas(ws)

    # Conditional formatting for Unit Expiration Date and Quarter columns
    table1_ref = get_table_ref(ws, "Table1") or table1_ref
    apply_unit_expiration_conditional_formatting(ws, table1_ref, unit_exp_col, quarter_col)

    # Column widths and alignment
    ws.column_dimensions["G"].width = 9
    ws.column_dimensions["K"].width = 9.17
    ws.column_dimensions["L"].width = 14
    for row in range(3, 100):
        ws.cell(row=row, column=9).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=12).alignment = Alignment(horizontal="center")

    # Date formatting for Unit Expiration Date and Registration Date
    last_row = find_last_row(ws, 1)
    for row in range(2, last_row + 1):
        ws.cell(row=row, column=unit_exp_col).number_format = "m/d/yy"
        ws.cell(row=row, column=reg_date_col).number_format = "m/d/yy"

    autosize_columns(ws)
    ws.column_dimensions["I"].width = 50

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def find_last_row(ws, col_idx: int) -> int:
    for row in range(ws.max_row, 0, -1):
        if ws.cell(row=row, column=col_idx).value not in (None, ""):
            return row
    return 0


def clear_tables(ws) -> None:
    if hasattr(ws, "_tables"):
        ws._tables.clear()
    else:
        ws.tables.clear()


def add_table(ws, name: str, ref: str, style_name: Optional[str] = None) -> None:
    table = Table(displayName=name, ref=ref)
    if style_name:
        table.tableStyleInfo = TableStyleInfo(
            name=style_name,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
    ws.add_table(table)


def update_table_ref(ws, name: str, ref: str) -> None:
    table = ws.tables.get(name)
    if table is None:
        return
    table.ref = ref
    if table.autoFilter:
        table.autoFilter.ref = ref


def get_table_ref(ws, name: str) -> Optional[str]:
    table = ws.tables.get(name)
    if table is None:
        return None
    return table.ref


def remove_duplicates_in_column(ws, col_idx: int, start_row: int, end_row: int) -> None:
    seen = set()
    write_row = start_row
    for row in range(start_row, end_row + 1):
        value = ws.cell(row=row, column=col_idx).value
        key = value if value is not None else ""
        if key in seen:
            continue
        seen.add(key)
        if write_row != row:
            ws.cell(row=write_row, column=col_idx).value = value
        write_row += 1
    for row in range(write_row, end_row + 1):
        ws.cell(row=row, column=col_idx).value = None


def count_values(ws, col_idx: int, start_row: int, end_row: int) -> dict:
    counts = {}
    for row in range(start_row, end_row + 1):
        value = ws.cell(row=row, column=col_idx).value
        counts[value] = counts.get(value, 0) + 1
    return counts


def normalize_date_column(ws, col_idx: int) -> None:
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if isinstance(cell.value, str) and "-" in cell.value:
            cell.value = cell.value.replace("-", "/")
        dt = parse_date(cell.value)
        if dt:
            cell.value = dt


def unmerge_row(ws, row_idx: int) -> None:
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        if merged_range.min_row <= row_idx <= merged_range.max_row:
            ws.unmerge_cells(str(merged_range))


def autosize_columns(ws) -> None:
    for col_idx, column_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            if isinstance(value, datetime):
                display = value.strftime("%m/%d/%y")
            else:
                display = str(value)
            max_len = max(max_len, len(display))
        if max_len > 0:
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def sort_table_by_column(ws, ref: str, sort_col: int, header_row: int, reverse: bool) -> None:
    ref_obj = parse_ref(ref)
    data_rows = []
    for row in range(header_row + 1, ref_obj.end_row + 1):
        row_values = [ws.cell(row=row, column=col).value for col in range(ref_obj.start_col, ref_obj.end_col + 1)]
        sort_value = ws.cell(row=row, column=sort_col).value
        data_rows.append((sort_value, row_values))

    data_rows.sort(key=lambda item: (item[0] is None, item[0]), reverse=reverse)

    for idx, (_, row_values) in enumerate(data_rows, start=header_row + 1):
        for offset, value in enumerate(row_values):
            ws.cell(row=idx, column=ref_obj.start_col + offset).value = value


def sort_table_by_date(ws, ref: str, date_col: int) -> None:
    ref_obj = parse_ref(ref)
    data_rows = []
    for row in range(ref_obj.start_row + 1, ref_obj.end_row + 1):
        row_values = [ws.cell(row=row, column=col).value for col in range(ref_obj.start_col, ref_obj.end_col + 1)]
        sort_value = parse_date(ws.cell(row=row, column=date_col).value)
        data_rows.append((sort_value, row_values))

    data_rows.sort(key=lambda item: (item[0] is None, item[0] or datetime.min))

    for idx, (_, row_values) in enumerate(data_rows, start=ref_obj.start_row + 1):
        for offset, value in enumerate(row_values):
            ws.cell(row=idx, column=ref_obj.start_col + offset).value = value


def sort_table_by_band_color(ws, ref: str, date_col: int) -> None:
    ref_obj = parse_ref(ref)
    current_year = datetime.now().year
    data_rows = []
    for row in range(ref_obj.start_row + 1, ref_obj.end_row + 1):
        row_values = [ws.cell(row=row, column=col).value for col in range(ref_obj.start_col, ref_obj.end_col + 1)]
        dt = parse_date(ws.cell(row=row, column=date_col).value)
        band = band_key_for_year(dt, current_year)
        data_rows.append((band, row_values))

    data_rows.sort(key=lambda item: item[0])

    for idx, (_, row_values) in enumerate(data_rows, start=ref_obj.start_row + 1):
        for offset, value in enumerate(row_values):
            ws.cell(row=idx, column=ref_obj.start_col + offset).value = value


def band_key_for_year(date_value: Optional[datetime], current_year: int) -> int:
    if date_value is None:
        return 0
    year = date_value.year
    if year < current_year:
        return 0
    if year >= current_year + 2:
        return 1
    if year == current_year + 1:
        return 2
    return 3


def date_to_quarter(value) -> str:
    dt = parse_date(value)
    if dt is None:
        return ""
    quarter = ((dt.month - 1) // 3) + 1
    return f"{dt.year} Q{quarter}"


def parse_date(value) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        try:
            return datetime.fromordinal(datetime(1899, 12, 30).toordinal() + int(value))
        except ValueError:
            return None
    if isinstance(value, str):
        for fmt in (
            "%m/%d/%Y",
            "%m/%d/%y",
            "%Y/%m/%d",
            "%Y-%m-%d",
            "%m/%d/%Y %H:%M:%S",
            "%Y-%m-%d %H:%M:%S",
        ):
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
    return None


def apply_unit_expiration_conditional_formatting(
    ws, table_ref: str, date_col: int, quarter_col: int
) -> None:
    tr = parse_ref(table_ref)
    start_row = tr.start_row + 1
    end_row = tr.end_row
    if end_row < start_row:
        return

    date_letter = get_column_letter(date_col)
    quarter_letter = get_column_letter(quarter_col)

    ref_cell = f"${date_letter}{start_row}"
    non_number_formula = f"=NOT(ISNUMBER({ref_cell}))"
    bad_formula = f"=AND(ISNUMBER({ref_cell}),YEAR({ref_cell})=YEAR(TODAY()))"
    neutral_formula = f"=AND(ISNUMBER({ref_cell}),YEAR({ref_cell})=YEAR(TODAY())+1)"
    good_formula = f"=AND(ISNUMBER({ref_cell}),YEAR({ref_cell})>=YEAR(TODAY())+2)"
    check_formula = f"=OR(NOT(ISNUMBER({ref_cell})),YEAR({ref_cell})<YEAR(TODAY()))"

    ws.conditional_formatting._cf_rules.clear()

    date_range = f"{date_letter}{start_row}:{date_letter}{end_row}"
    quarter_range = f"{quarter_letter}{start_row}:{quarter_letter}{end_row}"

    date_styles = build_style_set(
        bad_font="FF9C0006",
        neutral_font="FF9C5700",
        good_font="FF006100",
        check_font="FFFFFFFF",
    )
    quarter_styles = build_style_set(
        bad_font="FF9C0006",
        neutral_font="FF9C6500",
        good_font="FF006100",
        check_font="FFFFFFFF",
    )

    add_style_rules(
        ws,
        date_range,
        bad_formula,
        neutral_formula,
        good_formula,
        check_formula,
        date_styles,
        stop_if_true_check=True,
        stop_if_true_current=False,
        precheck_formula=non_number_formula,
    )
    add_style_rules(
        ws,
        quarter_range,
        bad_formula,
        neutral_formula,
        good_formula,
        check_formula,
        quarter_styles,
        stop_if_true_check=False,
        stop_if_true_current=True,
        precheck_formula=non_number_formula,
    )


def build_style_set(bad_font: str, neutral_font: str, good_font: str, check_font: str) -> dict:
    def solid_fill(color: str) -> PatternFill:
        return PatternFill(fill_type="solid", fgColor=color, bgColor=color)

    return {
        "bad": DifferentialStyle(
            fill=solid_fill("FFFFC7CE"),
            font=Font(color=bad_font),
            border=Border(),
        ),
        "neutral": DifferentialStyle(
            fill=solid_fill("FFFFEB9C"),
            font=Font(color=neutral_font),
            border=Border(),
        ),
        "good": DifferentialStyle(
            fill=solid_fill("FFC6EFCE"),
            font=Font(color=good_font),
            border=Border(),
        ),
        "check": DifferentialStyle(
            fill=solid_fill("FF7F7F7F"),
            font=Font(color=check_font),
            border=Border(),
        ),
    }


def add_style_rules(
    ws,
    cell_range: str,
    bad_formula: str,
    neutral_formula: str,
    good_formula: str,
    check_formula: str,
    styles: dict,
    stop_if_true_check: bool,
    stop_if_true_current: bool,
    precheck_formula: Optional[str] = None,
) -> None:
    rules = [
        Rule(type="expression", dxf=styles["check"], stopIfTrue=True, formula=[precheck_formula])
        if precheck_formula
        else None,
        Rule(type="expression", dxf=styles["bad"], stopIfTrue=stop_if_true_current, formula=[bad_formula]),
        Rule(type="expression", dxf=styles["neutral"], stopIfTrue=True, formula=[neutral_formula]),
        Rule(type="expression", dxf=styles["good"], stopIfTrue=True, formula=[good_formula]),
        Rule(type="expression", dxf=styles["check"], stopIfTrue=stop_if_true_check, formula=[check_formula]),
    ]
    for rule in rules:
        if rule is not None:
            ws.conditional_formatting.add(cell_range, rule)


def compute_quarter_counts(ws, quarter_col: int) -> List[Tuple[str, int]]:
    last_row = find_last_row(ws, quarter_col)
    counts = {}
    for row in range(2, last_row + 1):
        value = ws.cell(row=row, column=quarter_col).value
        if not value:
            continue
        counts[value] = counts.get(value, 0) + 1
    return sorted(counts.items())


def update_table2_count_formulas(ws) -> None:
    table_ref = get_table_ref(ws, "Table2")
    if not table_ref:
        return
    ref = parse_ref(table_ref)
    key_col_letter = get_column_letter(ref.start_col)
    for row in range(ref.start_row + 1, ref.end_row + 1):
        ws.cell(row=row, column=ref.end_col).value = f"=COUNTIF(B:B,{key_col_letter}{row})"


def write_quarter_counts(ws, counts: List[Tuple[str, int]], start_col: int, start_row: int) -> None:
    ws.cell(row=start_row, column=start_col).value = "Quarter"
    ws.cell(row=start_row, column=start_col + 1).value = "Count"
    for idx, (quarter, _count) in enumerate(counts, start=start_row + 1):
        ws.cell(row=idx, column=start_col).value = quarter
        ws.cell(row=idx, column=start_col + 1).value = (
            f"=COUNTIF(E:E,{get_column_letter(start_col)}{idx + 1})"
        )

    last_row = start_row + len(counts)
    table_ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(start_col + 1)}{last_row}"
    add_table(ws, "QuarterCounts", table_ref, style_name="TableStyleMedium12")
    table = ws.tables.get("QuarterCounts")
    if table:
        header_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(start_col + 1)}{start_row}"
        for row in ws[header_range]:
            for cell in row:
                cell.fill = PatternFill(fill_type=None)


def build_asset_report_header(ws) -> None:
    title_fill = PatternFill(fill_type="solid", fgColor="FF366092")
    green_fill = PatternFill(fill_type="solid", fgColor="FF9BBB59")
    white_font = Font(name="Calibri", size=18, color="FFFFFFFF", bold=True)
    header_font = Font(name="Calibri", size=18, color="FFFFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    double_bottom = Border(bottom=Side(style="double", color="FFFFFFFF"))

    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "Asset Report"
    cell.alignment = center
    cell.fill = title_fill
    cell.font = white_font

    ws.merge_cells("H1:I1")
    cell = ws["H1"]
    cell.value = "Asset Count"
    cell.alignment = center
    cell.style = "Accent2"
    cell.font = header_font
    cell.border = double_bottom

    ws.merge_cells("K1:L1")
    cell = ws["K1"]
    cell.value = "Renewal Schedule"
    cell.alignment = center
    cell.style = "Accent4"
    cell.font = header_font
    cell.border = double_bottom

    for addr in ("A1:F1", "H1:I1", "K1:L1"):
        for row in ws[addr]:
            for cell in row:
                cell.border = double_bottom

    for row in ws["A1:L1"]:
        for cell in row:
            if cell.value is not None:
                cell.font = white_font
                cell.alignment = center


def parse_ref(ref: str) -> TableRef:
    match = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", ref)
    if not match:
        raise ValueError(f"Invalid range: {ref}")
    start_col = column_index_from_string(match.group(1))
    start_row = int(match.group(2))
    end_col = column_index_from_string(match.group(3))
    end_row = int(match.group(4))
    return TableRef(start_col, start_row, end_col, end_row)


def expand_table_range(ref: str, insert_col: int) -> str:
    tr = parse_ref(ref)
    if tr.end_col >= insert_col:
        tr.end_col += 1
    if tr.start_col >= insert_col:
        tr.start_col += 1
    return f"{get_column_letter(tr.start_col)}{tr.start_row}:{get_column_letter(tr.end_col)}{tr.end_row}"


def shift_table_columns(ws, insert_col: int, offset: int) -> None:
    for table in ws.tables.values():
        tr = parse_ref(table.ref)
        if tr.start_col >= insert_col:
            tr.start_col += offset
        if tr.end_col >= insert_col:
            tr.end_col += offset
        table.ref = f"{get_column_letter(tr.start_col)}{tr.start_row}:{get_column_letter(tr.end_col)}{tr.end_row}"
        if table.autoFilter:
            table.autoFilter.ref = table.ref


def shift_table_rows(ws, insert_row: int, offset: int) -> None:
    for table in ws.tables.values():
        tr = parse_ref(table.ref)
        if tr.start_row >= insert_row:
            tr.start_row += offset
        if tr.end_row >= insert_row:
            tr.end_row += offset
        table.ref = f"{get_column_letter(tr.start_col)}{tr.start_row}:{get_column_letter(tr.end_col)}{tr.end_row}"
        if table.autoFilter:
            table.autoFilter.ref = table.ref


def find_header_column(ws, header_name: str) -> Optional[int]:
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == header_name:
            return col
    return None
