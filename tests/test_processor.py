import os
import sys
import unittest
from datetime import datetime

from openpyxl import Workbook

TESTS_DIR = os.path.dirname(__file__)
REPO_ROOT = os.path.abspath(os.path.join(TESTS_DIR, ".."))
sys.path.insert(0, os.path.join(REPO_ROOT, "backend"))

from app import processor


class TestDateHelpers(unittest.TestCase):
    def test_parse_date_accepts_datetime(self) -> None:
        dt = datetime(2024, 1, 2)
        self.assertEqual(processor.parse_date(dt), dt)

    def test_parse_date_accepts_excel_serial(self) -> None:
        value = 45000
        expected = datetime.fromordinal(datetime(1899, 12, 30).toordinal() + value)
        self.assertEqual(processor.parse_date(value), expected)

    def test_parse_date_accepts_string(self) -> None:
        dt = processor.parse_date("2024-03-15")
        self.assertIsNotNone(dt)
        self.assertEqual((dt.year, dt.month, dt.day), (2024, 3, 15))

    def test_parse_date_rejects_invalid(self) -> None:
        self.assertIsNone(processor.parse_date("not a date"))

    def test_date_to_quarter(self) -> None:
        self.assertEqual(processor.date_to_quarter("2024-05-20"), "2024 Q2")
        self.assertEqual(processor.date_to_quarter(None), "")


class TestTableRefHelpers(unittest.TestCase):
    def test_parse_ref(self) -> None:
        ref = processor.parse_ref("B2:D10")
        self.assertEqual((ref.start_col, ref.start_row, ref.end_col, ref.end_row), (2, 2, 4, 10))

    def test_expand_table_range_inserts_within(self) -> None:
        self.assertEqual(processor.expand_table_range("A1:E10", 3), "A1:F10")

    def test_expand_table_range_inserts_before(self) -> None:
        self.assertEqual(processor.expand_table_range("A1:E10", 1), "B1:F10")


class TestWorksheetTransforms(unittest.TestCase):
    def test_remove_duplicates_in_column(self) -> None:
        wb = Workbook()
        ws = wb.active
        values = ["Header", "A", "B", "A"]
        for idx, value in enumerate(values, start=1):
            ws.cell(row=idx, column=1).value = value

        processor.remove_duplicates_in_column(ws, 1, 1, 4)

        self.assertEqual(ws.cell(row=1, column=1).value, "Header")
        self.assertEqual(ws.cell(row=2, column=1).value, "A")
        self.assertEqual(ws.cell(row=3, column=1).value, "B")
        self.assertIsNone(ws.cell(row=4, column=1).value)

    def test_count_values(self) -> None:
        wb = Workbook()
        ws = wb.active
        values = ["Header", "A", "B", "A"]
        for idx, value in enumerate(values, start=1):
            ws.cell(row=idx, column=1).value = value

        counts = processor.count_values(ws, 1, 1, 4)
        self.assertEqual(counts["Header"], 1)
        self.assertEqual(counts["A"], 2)
        self.assertEqual(counts["B"], 1)

    def test_compute_quarter_counts(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "Quarter"
        ws.cell(row=2, column=1).value = "2024 Q1"
        ws.cell(row=3, column=1).value = ""
        ws.cell(row=4, column=1).value = "2024 Q1"
        ws.cell(row=5, column=1).value = "2025 Q3"

        counts = processor.compute_quarter_counts(ws, 1)
        self.assertEqual(counts, [("2024 Q1", 2), ("2025 Q3", 1)])

    def test_update_table2_count_formulas(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws["G1"] = "Asset"
        ws["H1"] = "Count"
        ws["G2"] = "A"
        ws["G3"] = "B"
        processor.add_table(ws, "Asset_Count", "G1:H3", style_name="TableStyleMedium10")

        processor.update_table2_count_formulas(ws)

        self.assertEqual(ws["H2"].value, "=COUNTIF(B:B,G2)")
        self.assertEqual(ws["H3"].value, "=COUNTIF(B:B,G3)")


if __name__ == "__main__":
    unittest.main()
